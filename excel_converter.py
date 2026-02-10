import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def convert_excel(input_path: str, output_path: str):
    # 데이터 로드
    df = pd.read_excel(input_path)
    df = df.drop(columns=['상품코드', '합포여부', '총개수', '위치'], errors='ignore')

    today_str = datetime.today().strftime("%y.%m.%d")

    # 스타일
    font_title = Font(size=24, bold=True)
    font_cell = Font(size=24)
    align_center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="CCF2FF")

    thin = Side(style="thin")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_fill = PatternFill("solid", fgColor="CCF2FF")

    def apply_box(ws, start_row, end_row):
        for r in range(start_row, end_row + 1):
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.border = box_border
                cell.alignment = align_center
                cell.font = font_cell

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for no_value in df["no"].dropna().unique():
            df_no = df[df["no"] == no_value].copy()
            output_round = str(df_no["출력차수"].iloc[0])

            # 개수 정리
            df_no["개수"] = pd.to_numeric(df_no["개수"], errors="coerce").fillna(0)

            # 🔹 상품명 + 옵션 기준 수량 합산
            df_no["수량"] = (
                df_no.groupby(["상품명", "옵션"])["개수"].transform("sum")
            )

            df_no = df_no.drop_duplicates(subset=["상품명", "옵션"])
            df_no = df_no.drop(columns=["개수"])

            df_no["loc_prefix"] = df_no["로케이션"].astype(str).str[0]
            df_no = df_no.sort_values(
                by=["loc_prefix", "로케이션", "상품명", "옵션"]
            )

            ws = writer.book.create_sheet(title=str(no_value))

            # 제목
            ws.merge_cells("A1:D1")
            red_font = InlineFont(color="FF0000", sz=24, b=True)
            black_font = InlineFont(color="000000", sz=24, b=True)

            ws["A1"].value = CellRichText(
                TextBlock(black_font, f"{today_str} "),
                TextBlock(red_font, f"{output_round}차 - {no_value} "),
                TextBlock(black_font, "에르모어 토탈피킹")
            )
            ws["A1"].alignment = align_center
            ws["A1"].border = box_border

            row_idx = 3
            total_qty = 0

            # 헤더
            headers = ["로케이션", "상품명", "옵션", "수량"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.fill = header_fill
                cell.font = font_cell
                cell.alignment = align_center
                cell.border = box_border

            row_idx += 1

            for _, r in df_no.iterrows():
                ws.cell(row_idx, 1, r["로케이션"])
                ws.cell(row_idx, 2, r["상품명"])
                ws.cell(row_idx, 3, r["옵션"])
                ws.cell(row_idx, 4, int(r["수량"]))

                for c in range(1, 5):
                    ws.cell(row_idx, c).border = box_border
                    ws.cell(row_idx, c).alignment = align_center
                    ws.cell(row_idx, c).font = font_cell

                total_qty += int(r["수량"])
                row_idx += 1

            # 총합
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws.cell(row_idx, 1, "총합계")
            ws.cell(row_idx, 4, total_qty)

            for c in range(1, 5):
                cell = ws.cell(row_idx, c)
                cell.fill = total_fill
                cell.font = font_title
                cell.alignment = align_center
                cell.border = box_border

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 80
            ws.column_dimensions["C"].width = 80
            ws.column_dimensions["D"].width = 14

    return output_path